import json
import openpyxl
import datetime

from python_files.constants import EXCEL_2024_COORDINATES

# CHANGES FOR THIS FILE: Change EXCEL_2024_COORDINATES to EXCEL_[NEW_CURRENT_YEAR]_COORDINATES
def create_new_hda_file(company):
    with open('json_files/json_data.json', 'r') as file:
        data = json.load(file)

    template_path = 'excel_sheets/Bayshore_Template.xlsm' if company == 'Bayshore' else 'excel_sheets/Unichem_Template.xlsm'
    workbook = openpyxl.load_workbook(template_path, keep_vba=True)
    copied_sheet = workbook.active

    skip = ['item_packing', 'gtin_14']
    for k, v in EXCEL_2024_COORDINATES.items():
        if k in skip:   # skip item packing and gtin because they have separate implementations
            continue
        if k == 'country_of_origin' and company == "UNICHEM":       # country of origin never changes for Unichem
            continue
        r, c = v
        copied_sheet.cell(row=r, column=c).value = data[k]

    row_item_packing_start, col_item_packing_start = EXCEL_2024_COORDINATES['item_packing']
    row_gtin_start, col_gtin_start = EXCEL_2024_COORDINATES['gtin_14']

    populate_item_packing(copied_sheet, row_item_packing_start, col_item_packing_start, data)
    populate_gtin_14(copied_sheet, row_gtin_start, col_gtin_start, data)
    populate_remaining_fields(copied_sheet, data)

    # code for parsing for title of Excel sheet
    drug_type = 'Tabs' if data['dosage_form'] == 'Tablet' else 'Capsules'

    if company == "UNICHEM":
        new_title = f"{data['description'].split(' ')[0]} {drug_type}, {data['strength']} {data['size'].replace('CT', '')} CT HDA - New Form 2024"
    else:
        new_title = f"{data['description'].split(' ')[0]} - HDA {data['ndc']} Revised {data['todays_date'].replace('/','.')}"

    workbook.save(f"excel_sheets/{new_title}.xlsm")

    return f"excel_sheets/{new_title}.xlsm"


def populate_item_packing(copied_sheet, row_start, col_start, data):
    r = 0
    for row in range(row_start, row_start + 8, 2):
        c = 0
        for col in range(col_start, col_start + 6):
            copied_sheet.cell(row, col).value = data['item_packing'][r][c]
            c += 1
        r += 1


def populate_gtin_14(copied_sheet, r, c, data):
    for i in range(3):
        copied_sheet.cell(r + i, c).value = data['gtin_14'][i]


# hard-coded data fields
def populate_remaining_fields(copied_sheet, data):
    # Make edits here to hard-coded coordinates with new HDA form coordinates
    ##################################################################
    copied_sheet.cell(43, 29).value = data['inner_packet_quantity']
    copied_sheet.cell(70, 19).value = data['gtin_14'][0]
    copied_sheet.cell(4, 31).value = datetime.date.today().strftime("%m/%d/%Y")
    ##################################################################
