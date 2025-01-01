from python_files.constants import EXCEL_2014_COORDINATES, EXCEL_2017_COORDINATES, EXCEL_2021_COORDINATES, \
    EXCEL_2024_COORDINATES
import json
import openpyxl
import os
from datetime import datetime

# reads data fields from uploaded Excel sheet
def read_hda_sheet(file_name):
    if not os.path.exists('excel_sheets/' + file_name):
        print("File does not exist")
        return
    workbook = openpyxl.load_workbook('excel_sheets/' + file_name)
    sheet = workbook.active

    with open('json_files/configurations.json', 'r') as file:
        configurations = json.load(file)
    file.close()
    with open('json_files/36Months.json', 'r') as file:
        thirty_six_months = json.load(file)
    file.close()
    with open('json_files/pricing.json', 'r') as file:
        pricing = json.load(file)
    file.close()

    # import data dictionary in the constant Python file. If there is additional data for the year 2025,
    # there needs to be a if-statement below
    coordinates = EXCEL_2014_COORDINATES
    try:
        if '2017' in sheet.cell(row=4, column=3).value:
            coordinates = EXCEL_2017_COORDINATES
        elif '2021' in sheet.cell(row=4, column=3).value:
            coordinates = EXCEL_2021_COORDINATES
        elif '2024' in sheet.cell(row=4, column=3).value:
            coordinates = EXCEL_2024_COORDINATES

        # Add logic here for years beyond 2024
        ##################################################################

        ##################################################################
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        hda_product_data = {}

        # reads in key-value (product_name, coordinate) pairs from constants.py for respective year
        for k, v in coordinates.items():
            # writes to hda_product_data dictionary by mapping a product_name to its product_data value
            # these values are obtained by reading directly from the coordinates in the uploaded Excel sheet
            hda_product_data[k] = sheet.cell(row=v[0], column=v[1]).value

        # For better convenience, in the dictionary, item_packing data is stored as a 4x6 2D array
        row_item_packing_start, col_item_packing_start = coordinates['item_packing']
        hda_product_data['item_packing'] = populate_item_packing(sheet, row_item_packing_start, col_item_packing_start)

        # For better convenience, in the dictionary, gtin data is stored as a 1x3 1D array
        gtin_row_start, gtin_col_start = coordinates['gtin_14']
        hda_product_data['gtin_14'] = populate_gtin(sheet, gtin_row_start, gtin_col_start)

        # these data fields are read from json files because the data was needed to be double checked in these files
        ndc = hda_product_data['ndc'].replace('-', '')
        hda_product_data['inner_packet_quantity'] = configurations[ndc]['Inner Case Qty (ea)']
        hda_product_data['case_quantity'] = configurations[ndc]['Outer Case Qty (ea)']
        hda_product_data['regular_cost'] = pricing[ndc]['Regular Cost (AWP)']
        hda_product_data['invoice_cost'] = pricing[ndc]['Invoice Cost (WAC)']
        hda_product_data['shelf_life'] = '36' if ndc in thirty_six_months else '24'
        # hda_product_data['todays_date'] = datetime.date.today().strftime("%m/%d/%Y")
        if hda_product_data['as_of_date']:
            if isinstance(hda_product_data['as_of_date'], datetime):
                hda_product_data['as_of_date'] = hda_product_data['as_of_date'].strftime("%m/%d/%Y")
        else:
            hda_product_data['as_of_date'] = ''

        hda_product_data['description'] = hda_product_data['description'].replace('/', '-')
        hda_product_data['strength'] = hda_product_data['strength'].replace('/', '-')

        with open("json_files/json_data.json", "w") as outfile:
            json.dump(hda_product_data, outfile, indent=4, sort_keys=True, default=str)


'''
To populate item packing, the only coordinate needed is the coordinate of the left_most upper_most tile in the item packing
    table in the old HDA sheet.

In this method, given a starting row and col, it reads traverses through the item packing table, which is of size 4x6
'''


def populate_item_packing(sheet, row_start, col_start):
    item_packing = [['' for _ in range(6)] for _ in range(4)]
    r = 0
    for row in range(row_start, row_start + 8, 2):  # the coordinate for the rows increment by 2
        c = 0
        for col in range(col_start, col_start + 6):  # the coordinates for the columns increment by 1
            item_packing[r][c] = sheet.cell(row=row, column=col).value
            c += 1
        r += 1
    return item_packing


'''
Same logic with item packing. Gtin data is 1x3 size array
'''


def populate_gtin(sheet, row_start, col_start):
    gtin_14 = [''] * 3
    for i in range(3):
        gtin_14[i] = str(sheet.cell(row=row_start + i, column=col_start).value)
    gtin_14[0] = '00' + gtin_14[0] if len(gtin_14[0]) != 14 else gtin_14[0]
    return gtin_14
