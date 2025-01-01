from python_files.constants import EXCEL_2014_COORDINATES, EXCEL_2017_COORDINATES, EXCEL_2021_COORDINATES
from python_files.constants import BAYSHORE_2021_COORDINATES
import json
import openpyxl
import datetime
import os


def read_bayshore_sheet(file_name):
    if not os.path.exists('excel_sheets/' + file_name):
        print("File does not exist")
        return
    workbook = openpyxl.load_workbook('excel_sheets/' + file_name)
    sheet = workbook.active

    coordinates = EXCEL_2014_COORDINATES
    bayshore_coordinates = BAYSHORE_2021_COORDINATES
    try:
        if '2017' in sheet.cell(row=4, column=3).value:
            coordinates = EXCEL_2017_COORDINATES
        elif '2021' in sheet.cell(row=4, column=3).value:
            coordinates = EXCEL_2021_COORDINATES
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        hda_product_data = {}

        for k, v in coordinates.items():
            hda_product_data[k] = sheet.cell(row=v[0], column=v[1]).value

        row_item_packing_start, col_item_packing_start = coordinates['item_packing']
        hda_product_data['item_packing'] = populate_item_packing(sheet, row_item_packing_start, col_item_packing_start)

        gtin_row_start, gtin_col_start = coordinates['gtin_14']
        hda_product_data['gtin_14'] = populate_gtin(sheet, gtin_row_start, gtin_col_start)

        for k, v in bayshore_coordinates.items():
            hda_product_data[k] = sheet.cell(row=v[0], column=v[1]).value

        hda_product_data['todays_date'] = datetime.date.today().strftime("%m/%d/%Y")

        if hda_product_data['as_of_date']:
            hda_product_data['as_of_date'] = hda_product_data['as_of_date'].strftime("%m/%d/%Y")

        hda_product_data['description'] = hda_product_data['description'].replace('/', '-')
        hda_product_data['strength'] = hda_product_data['strength'].replace('/', '-')

        with open("json_files/json_data.json", "w") as outfile:
            json.dump(hda_product_data, outfile, indent=4, sort_keys=True, default=str)


def populate_item_packing(sheet, row_start, col_start):
    item_packing = [['' for _ in range(6)] for _ in range(4)]
    r = 0
    for row in range(row_start, row_start + 8, 2):
        c = 0
        for col in range(col_start, col_start + 6):
            item_packing[r][c] = sheet.cell(row=row, column=col).value
            c += 1
        r += 1

    return item_packing


def populate_gtin(sheet, row_start, col_start):
    gtin_14 = [''] * 3
    for i in range(3):
        gtin_14[i] = str(sheet.cell(row=row_start + i, column=col_start).value)
    return gtin_14

